from datetime import datetime
import io
import re
import unicodedata

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from modules.database import connect_db


_TIME_HHMM_PATTERN = re.compile(r"^\d{2}:\d{2}$")
VALID_STATES = {"pendiente", "en_progreso", "completado"}

DEFAULT_ACTIVITY_CATALOG: dict[str, dict[str, object]] = {
    "taller_reparacion": {
        "label": "Taller/Reparacion",
        "project_options": [
            "Diagnostico",
            "Reparacion mecanica",
            "Reparacion electronica",
            "Mantenimiento preventivo",
            "Mantenimiento correctivo",
            "Calibracion",
            "Ensamblaje/montaje",
            "Pruebas de funcionamiento",
            "Entrega al cliente",
        ],
    },
    "programacion": {
        "label": "Programacion",
        "project_options": [
            "Desarrollo de software",
            "Correccion de errores",
            "Automatizacion",
            "Programacion de microcontroladores",
            "Configurar sistemas",
            "Pruebas",
            "Documentacion tecnica",
        ],
    },
    "diseno": {
        "label": "Diseno",
        "project_options": [
            "Diseno CAD",
            "Modelado 3D",
            "Diseno de PCB",
            "Diseno de circuitos",
            "Diseno de interfaces",
            "Renderizado",
            "Diseno grafico",
        ],
    },
    "administracion": {
        "label": "Administracion",
        "project_options": [
            "Atencion al cliente",
            "Cotizaciones",
            "Compra de materiales",
            "Inventariado",
            "Capacitaciones",
            "Reuniones",
            "Organizacion del taller",
        ],
    },
    "produccion_fabricacion": {
        "label": "Produccion/Fabricacion",
        "project_options": [
            "Corte",
            "Impresion 3D",
            "CNC/Laser",
            "Pintura",
            "Acabado",
            "Empaquetado",
        ],
    },
}

DEFAULT_STATE_OPTIONS = [
    {"value": "pendiente", "label": "Pendiente"},
    {"value": "en_progreso", "label": "En progreso"},
    {"value": "completado", "label": "Completado"},
]


def _slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    lowered = ascii_only.lower().strip()
    lowered = re.sub(r"[^a-z0-9]+", "_", lowered)
    lowered = re.sub(r"_+", "_", lowered).strip("_")
    return lowered


def get_activity_catalog() -> dict[str, dict[str, object]]:
    with connect_db() as conn:
        categories = conn.execute(
            "SELECT category_key, label FROM activity_categories ORDER BY label ASC"
        ).fetchall()
        options = conn.execute(
            "SELECT category_key, option_name FROM activity_options ORDER BY option_name ASC"
        ).fetchall()

    if not categories:
        return DEFAULT_ACTIVITY_CATALOG

    option_map: dict[str, list[str]] = {}
    for row in options:
        option_map.setdefault(row["category_key"], []).append(row["option_name"])

    catalog: dict[str, dict[str, object]] = {}
    for category in categories:
        catalog[category["category_key"]] = {
            "label": category["label"],
            "project_options": option_map.get(category["category_key"], []),
        }
    return catalog


def get_state_options() -> list[dict[str, str]]:
    with connect_db() as conn:
        rows = conn.execute(
            "SELECT state_value, label FROM record_states ORDER BY label ASC"
        ).fetchall()

    if not rows:
        return DEFAULT_STATE_OPTIONS

    return [{"value": row["state_value"], "label": row["label"]} for row in rows]


def add_category(label: str) -> tuple[bool, str]:
    cleaned_label = label.strip()
    if len(cleaned_label) < 2:
        return False, "La categoria debe tener al menos 2 caracteres"

    category_key = _slugify(cleaned_label)
    if not category_key:
        return False, "La categoria no tiene un formato valido"

    with connect_db() as conn:
        existing = conn.execute(
            "SELECT category_key FROM activity_categories WHERE category_key = ? OR label = ?",
            (category_key, cleaned_label),
        ).fetchone()
        if existing:
            return False, "La categoria ya existe"
        conn.execute(
            "INSERT INTO activity_categories (category_key, label) VALUES (?, ?)",
            (category_key, cleaned_label),
        )
    return True, "Categoria creada"


def add_activity_option(category_key: str, option_name: str) -> tuple[bool, str]:
    cleaned_category = category_key.strip()
    cleaned_option = option_name.strip()
    if not cleaned_category:
        return False, "Debes seleccionar una categoria"
    if len(cleaned_option) < 2:
        return False, "La actividad debe tener al menos 2 caracteres"

    with connect_db() as conn:
        category_exists = conn.execute(
            "SELECT 1 FROM activity_categories WHERE category_key = ?",
            (cleaned_category,),
        ).fetchone()
        if not category_exists:
            return False, "La categoria seleccionada no existe"

        exists = conn.execute(
            "SELECT 1 FROM activity_options WHERE category_key = ? AND option_name = ?",
            (cleaned_category, cleaned_option),
        ).fetchone()
        if exists:
            return False, "La actividad ya existe en esa categoria"

        conn.execute(
            "INSERT INTO activity_options (category_key, option_name) VALUES (?, ?)",
            (cleaned_category, cleaned_option),
        )
    return True, "Actividad agregada"


def add_state(label: str) -> tuple[bool, str]:
    cleaned_label = label.strip()
    if len(cleaned_label) < 2:
        return False, "El estado debe tener al menos 2 caracteres"

    state_value = _slugify(cleaned_label)
    if not state_value:
        return False, "El estado no tiene un formato valido"

    with connect_db() as conn:
        exists = conn.execute(
            "SELECT 1 FROM record_states WHERE state_value = ? OR label = ?",
            (state_value, cleaned_label),
        ).fetchone()
        if exists:
            return False, "Ese estado ya existe"

        conn.execute(
            "INSERT INTO record_states (state_value, label) VALUES (?, ?)",
            (state_value, cleaned_label),
        )
    return True, "Estado agregado"


def delete_category(category_key: str) -> tuple[bool, str]:
    cleaned_key = category_key.strip()
    if not cleaned_key:
        return False, "Categoria invalida"

    with connect_db() as conn:
        exists = conn.execute(
            "SELECT 1 FROM activity_categories WHERE category_key = ?",
            (cleaned_key,),
        ).fetchone()
        if not exists:
            return False, "La categoria no existe"

        in_use = conn.execute(
            "SELECT COUNT(*) AS total FROM registros WHERE categoria = ?",
            (cleaned_key,),
        ).fetchone()["total"]
        if int(in_use) > 0:
            return False, "No se puede eliminar una categoria con registros asociados"

        conn.execute("DELETE FROM activity_categories WHERE category_key = ?", (cleaned_key,))
    return True, "Categoria eliminada"


def delete_activity_option(category_key: str, option_name: str) -> tuple[bool, str]:
    cleaned_key = category_key.strip()
    cleaned_option = option_name.strip()
    if not cleaned_key or not cleaned_option:
        return False, "Datos de actividad invalidos"

    with connect_db() as conn:
        exists = conn.execute(
            "SELECT 1 FROM activity_options WHERE category_key = ? AND option_name = ?",
            (cleaned_key, cleaned_option),
        ).fetchone()
        if not exists:
            return False, "La actividad no existe"

        in_use = conn.execute(
            """
            SELECT COUNT(*) AS total
            FROM registros
            WHERE categoria = ?
              AND (proyecto_equipo = ? OR actividad = ?)
            """,
            (cleaned_key, cleaned_option, cleaned_option),
        ).fetchone()["total"]
        if int(in_use) > 0:
            return False, "No se puede eliminar una actividad con registros asociados"

        conn.execute(
            "DELETE FROM activity_options WHERE category_key = ? AND option_name = ?",
            (cleaned_key, cleaned_option),
        )
    return True, "Actividad eliminada"


def delete_state(state_value: str) -> tuple[bool, str]:
    cleaned_value = state_value.strip()
    if not cleaned_value:
        return False, "Estado invalido"
    if cleaned_value in {"pendiente", "en_progreso", "completado"}:
        return False, "No se puede eliminar un estado base del sistema"

    with connect_db() as conn:
        exists = conn.execute(
            "SELECT 1 FROM record_states WHERE state_value = ?",
            (cleaned_value,),
        ).fetchone()
        if not exists:
            return False, "El estado no existe"

        in_use = conn.execute(
            "SELECT COUNT(*) AS total FROM registros WHERE estado = ?",
            (cleaned_value,),
        ).fetchone()["total"]
        if int(in_use) > 0:
            return False, "No se puede eliminar un estado con registros asociados"

        conn.execute("DELETE FROM record_states WHERE state_value = ?", (cleaned_value,))
    return True, "Estado eliminado"


def _parse_hhmm(value: str, field_name: str) -> datetime:
    raw_value = value.strip()
    if not _TIME_HHMM_PATTERN.fullmatch(raw_value):
        raise ValueError(f"{field_name} debe tener el formato HH:MM")

    try:
        return datetime.strptime(raw_value, "%H:%M")
    except ValueError as exc:
        raise ValueError(f"{field_name} no es una hora valida") from exc


def calculate_worked_hours(hora_inicio: str, hora_fin: str, precision: int = 2) -> float:
    start = _parse_hhmm(hora_inicio, "hora_inicio")
    end = _parse_hhmm(hora_fin, "hora_fin")

    if end <= start:
        raise ValueError("La hora de fin debe ser mayor que la hora de inicio")

    diff_hours = (end - start).total_seconds() / 3600
    return round(diff_hours, precision)


def calculate_hours(hora_inicio: str, hora_fin: str) -> float:
    return calculate_worked_hours(hora_inicio, hora_fin)


_ESTADO_LABELS = {
    "pendiente": "Pendiente",
    "en_progreso": "En progreso",
    "completado": "Completado",
}

_HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FILL = PatternFill("solid", fgColor="DBEAFE")
_TOTAL_FONT = Font(bold=True)


def build_excel(registros, total_horas: float) -> bytes:
    catalog = get_activity_catalog()
    rows = [
        {
            "Fecha": r["fecha"],
            "Fecha estimada": r["fecha_estimada"],
            "Categoria": catalog.get(r["categoria"], {}).get("label", r["categoria"]),
            "Proyecto/Equipo": r["proyecto_equipo"],
            "Cliente/Codigo/Proyecto": r["cliente_referencia"],
            "Actividad": r["actividad"],
            "Estado": _ESTADO_LABELS.get(r["estado"], r["estado"]),
            "Descripcion": r["descripcion"],
            "Hora inicio": r["hora_inicio"],
            "Hora fin": r["hora_fin"],
            "Horas totales": round(float(r["horas_totales"]), 2),
        }
        for r in registros
    ]

    df = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Registros", index=False, startrow=0)
        ws = writer.sheets["Registros"]

        for col_idx, _ in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

        total_row = len(df) + 2
        total_label_col = len(df.columns) - 1
        total_value_col = len(df.columns)
        ws.cell(row=total_row, column=total_label_col).value = "TOTAL"
        ws.cell(row=total_row, column=total_label_col).font = _TOTAL_FONT
        ws.cell(row=total_row, column=total_value_col).value = round(total_horas, 2)
        ws.cell(row=total_row, column=total_value_col).font = _TOTAL_FONT
        for col_idx in range(total_label_col, total_value_col + 1):
            ws.cell(row=total_row, column=col_idx).fill = _TOTAL_FILL

    return output.getvalue()


def normalize_form_data(form_data: dict[str, str]) -> dict[str, str | float | None]:
    catalog = get_activity_catalog()
    valid_states = {state["value"] for state in get_state_options()}

    for field in [
        "fecha",
        "categoria",
        "proyecto_equipo",
        "hora_inicio",
    ]:
        if not form_data.get(field):
            raise ValueError(f"El campo '{field}' es obligatorio")

    categoria = form_data.get("categoria", "").strip()
    if categoria not in catalog:
        raise ValueError("La categoria seleccionada no es valida")

    proyecto_equipo = form_data.get("proyecto_equipo", "").strip()
    valid_project_options = catalog[categoria]["project_options"]
    if proyecto_equipo not in valid_project_options:
        raise ValueError("La actividad seleccionada no es valida para la categoria")

    cliente_referencia = form_data.get("cliente_referencia", "").strip()
    fecha_estimada = form_data.get("fecha_estimada", "").strip() or None

    hora_fin_raw = form_data.get("hora_fin", "").strip()
    is_open_task = not hora_fin_raw

    estado = form_data.get("estado", "pendiente").strip().lower()
    if is_open_task:
        estado = "en_progreso"
    elif estado not in valid_states:
        raise ValueError("El estado no es valido")

    try:
        datetime.strptime(form_data["fecha"], "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError("La fecha debe tener el formato YYYY-MM-DD") from exc

    try:
        datetime.strptime(form_data["hora_inicio"], "%H:%M")
    except ValueError as exc:
        raise ValueError("Las horas deben tener el formato HH:MM") from exc

    hours = 0.0
    if not is_open_task:
        try:
            datetime.strptime(hora_fin_raw, "%H:%M")
        except ValueError as exc:
            raise ValueError("Las horas deben tener el formato HH:MM") from exc

        hours = calculate_hours(form_data["hora_inicio"], hora_fin_raw)
    return {
        "fecha": form_data["fecha"],
        "fecha_estimada": fecha_estimada,
        "categoria": categoria,
        "proyecto_equipo": proyecto_equipo,
        "cliente_referencia": cliente_referencia,
        "actividad": proyecto_equipo,
        "estado": estado,
        "descripcion": form_data.get("descripcion", "").strip(),
        "hora_inicio": form_data["hora_inicio"],
        "hora_fin": hora_fin_raw,
        "horas_totales": hours,
    }
